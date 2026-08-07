"""
작업 대기열 API 엔드포인트
"""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, WebSocket, WebSocketDisconnect
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import os
import asyncio
import json
import logging
import redis.asyncio as aioredis

from app.database import get_db
from app import models
from pydantic import BaseModel

logger = logging.getLogger(__name__)

router = APIRouter(tags=["work_queue"])


# === Pydantic Schemas ===

class WorkQueueItemCreate(BaseModel):
    title: str
    description: Optional[str] = None
    hashtags: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    video_file_path: str
    source_type: Optional[str] = "MANUAL"
    approval_required: bool = False
    upload_method: Optional[str] = "API"
    target_platforms: Optional[List[str]] = ["youtube"]
    platform_configs: Optional[dict] = None
    scheduled_upload_time: Optional[datetime] = None
    enable_shopping_tag: bool = False
    shopping_tag_keyword: Optional[str] = None


class WorkQueueItemUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    hashtags: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    video_file_path: Optional[str] = None
    source_external_id: Optional[str] = None
    approval_status: Optional[str] = None
    upload_method: Optional[str] = None
    target_platforms: Optional[List[str]] = None
    platform_configs: Optional[dict] = None
    enable_shopping_tag: Optional[bool] = None
    shopping_tag_keyword: Optional[str] = None
    status: Optional[str] = None
    scheduled_upload_time: Optional[datetime] = None


class WorkQueueItemResponse(BaseModel):
    id: int
    title: str
    description: Optional[str] = None
    hashtags: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    video_file_path: Optional[str] = None
    # Source & Quality
    source_type: Optional[str] = None
    source_batch_id: Optional[str] = None
    source_external_id: Optional[str] = None
    approval_required: bool
    approval_status: str
    rejection_reason: Optional[str] = None
    # Upload Config
    upload_method: Optional[str] = None
    target_platforms: Optional[List[str]] = None
    platform_configs: Optional[dict] = None
    upload_priority: int
    scheduled_upload_time: Optional[datetime] = None  # [NEW]
    # Status
    status: str
    upload_progress: int
    uploaded_urls: Optional[dict] = None
    failure_reason: Optional[str] = None
    # Timestamps
    created_at: datetime
    updated_at: Optional[datetime] = None
    
    class Config:
        from_attributes = True


class BatchApproveRequest(BaseModel):
    item_ids: List[int]
    approved_by: str = "system"


class BatchRejectRequest(BaseModel):
    item_ids: List[int]
    reason: str


class BatchDeleteRequest(BaseModel):
    item_ids: List[int]

class BatchShieldConfigRequest(BaseModel):
    item_ids: List[int]
    shield_configs: dict


class PriorityUpdateRequest(BaseModel):
    item_id: int
    priority: int


class BatchResetRequest(BaseModel):
    item_ids: List[int]


class ExpertApprovalRequest(BaseModel):
    script: Optional[str] = None
    instructions: Optional[str] = None
    update_master_identity: bool = False
    approved_by: str = "expert"


class KeywordExtractionRequest(BaseModel):
    title: str
    description: str


class DraftItemCreate(BaseModel):
    """임시 등록 - 기본 정보만 저장 (제목, 설명, 태그) -- 영상은 나중에 첨부"""
    title: str
    description: Optional[str] = None
    hashtags: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    source_type: Optional[str] = "BULK_IMPORT"
    upload_method: Optional[str] = "BROWSER_AUTO"
    target_platforms: Optional[List[str]] = ["youtube"]
    platform_configs: Optional[dict] = None
    scheduled_upload_time: Optional[datetime] = None
    source_batch_id: Optional[str] = None
    source_external_id: Optional[str] = None  # JSON/Excel 각 행과 영상의 연결고리
    source_metadata: Optional[dict] = None


class AttachVideoRequest(BaseModel):
    """Draft에 영상 경로를 첨부 (external_id로 Draft 식별 가능)"""
    video_file_path: str
    source_external_id: Optional[str] = None  # ID 대신 external_id로 Draft 찾기 지원


class FinalizeRequest(BaseModel):
    """최종 등록 -- PENDING->QUEUED 전환. 승인 필요 여부 지정"""
    approval_required: bool = False
    scheduled_upload_time: Optional[datetime] = None
    upload_method: Optional[str] = None
    target_platforms: Optional[List[str]] = None


class BulkImportRequest(BaseModel):
    """Bulk import from JSON array or CSV data"""
    items: List[DraftItemCreate]
    source_batch_id: Optional[str] = None


# === API Endpoints ===

@router.post("/extract-shopping-keyword")
def extract_shopping_keyword_api(
    req: KeywordExtractionRequest,
    db: Session = Depends(get_db)
):
    """AI를 이용해 제목과 설명에서 쇼핑 키워드를 추출합니다."""
    settings = db.query(models.Settings).first()
    if not settings:
        raise HTTPException(500, "Settings not found")
        
    from app.services.ai_analyzer import ContentAnalyzer
    analyzer = ContentAnalyzer(db, settings)
    
    try:
        keyword = analyzer.extract_shopping_keyword(req.title, req.description)
        if keyword.upper() == "NONE":
            return {"keyword": ""}
        return {"keyword": keyword}
    except Exception as e:
        logger.error(f"Failed to extract shopping keyword: {e}")
        raise HTTPException(500, f"Keyword extraction failed: {str(e)}")


@router.get("/items", response_model=List[WorkQueueItemResponse])
def get_queue_items(
    status: Optional[str] = None,
    approval_status: Optional[str] = None,
    date_filter: Optional[str] = None, # "today", "week", "month", "all"
    source_batch_id: Optional[str] = None,
    source_external_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """작업 대기열 목록 조회"""
    query = db.query(models.WorkQueueItem)
    
    if status:
        query = query.filter(models.WorkQueueItem.status == status)
    
    if approval_status:
        query = query.filter(models.WorkQueueItem.approval_status == approval_status)
    
    if source_batch_id:
        query = query.filter(models.WorkQueueItem.source_batch_id == source_batch_id)
    
    if source_external_id:
        query = query.filter(models.WorkQueueItem.source_external_id == source_external_id)
        
    if date_filter and date_filter != "all":
        now = datetime.now()
        if date_filter == "today":
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_filter == "week":
            start_date = now - __import__('datetime').timedelta(days=7)
        elif date_filter == "month":
            start_date = now - __import__('datetime').timedelta(days=30)
        else:
            start_date = None
            
        if start_date:
            query = query.filter(models.WorkQueueItem.created_at >= start_date)
    
    items = query.order_by(
        models.WorkQueueItem.upload_priority.desc(),
        models.WorkQueueItem.created_at.desc()
    ).offset(skip).limit(limit).all()
    
    return items


# ... imports ...
from app.services.native_queue_worker import native_worker

# ... (get_queue_items remains same)

@router.post("/items", response_model=WorkQueueItemResponse)
def create_queue_item(
    item_data: WorkQueueItemCreate,
    db: Session = Depends(get_db)
):
    """작업 대기열에 항목 추가 (Auto-Approve 지원)"""
    
    # DISCOVERY 타입은 YouTube URL을 직접 경로로 사용 (로컬 파일 없음)
    is_discovery = (item_data.source_type or '').upper() == 'DISCOVERY'
    
    if not is_discovery:
        # 파일 존재 확인 (로컬 파일인 경우만)
        if not os.path.exists(item_data.video_file_path):
            raise HTTPException(404, f"Video file not found: {item_data.video_file_path}")
            
        # [NEW] 파일 안전 복사 로직
        settings = db.query(models.Settings).first()
        safe_dir = os.path.join(settings.root_download_path if settings and settings.root_download_path else os.getcwd(), "work_queue_uploads")
        os.makedirs(safe_dir, exist_ok=True)
        
        # 고유한 파일명 생성
        import shutil
        import uuid
        ext = os.path.splitext(item_data.video_file_path)[1]
        safe_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
        safe_file_path = os.path.join(safe_dir, safe_filename)
        
        try:
            shutil.copy2(item_data.video_file_path, safe_file_path)
            logger.info(f"📁 Video safely copied to: {safe_file_path}")
        except Exception as e:
            logger.error(f"Failed to copy video file: {e}")
            safe_file_path = item_data.video_file_path
    else:
        # DISCOVERY: YouTube URL을 그대로 경로로 저장
        safe_file_path = item_data.video_file_path
        logger.info(f"[SEARCH] Discovery item, using URL as path: {safe_file_path}")
    
    # Determine initial status based on approval_required
    initial_status = "QUEUED"
    if item_data.scheduled_upload_time and item_data.scheduled_upload_time > datetime.now():
        initial_status = "SCHEDULED_UPLOAD"
        logger.info(f"📅 Item scheduled for {item_data.scheduled_upload_time}")
    
    if item_data.approval_required:
        initial_approval = "PENDING"
    else:
        # Auto-Approve
        initial_approval = "AUTO_APPROVED"
    
    # WorkQueueItem 생성
    queue_item = models.WorkQueueItem(
        title=item_data.title,
        description=item_data.description,
        hashtags=item_data.hashtags,
        tags=item_data.tags,
        video_file_path=safe_file_path, # [UPDATED] Use safe copy
        source_type=item_data.source_type,
        approval_required=item_data.approval_required,
        approval_status=initial_approval,
        # Upload Config
        upload_method=item_data.upload_method,
        target_platforms=item_data.target_platforms,
        platform_configs=item_data.platform_configs or {},
        scheduled_upload_time=item_data.scheduled_upload_time, # [NEW]
        enable_shopping_tag=item_data.enable_shopping_tag,
        shopping_tag_keyword=item_data.shopping_tag_keyword,
        # Initial State
        status=initial_status,
        upload_progress=0,
        created_at=datetime.now()
    )
    
    db.add(queue_item)
    db.flush()  # ID 생성을 위해 flush
    
    # 규칙 엔진 적용
    from app.services.rule_engine import RuleEngine
    rule_engine = RuleEngine(db)
    
    actions = rule_engine.evaluate_rules(queue_item)
    if actions:
        logger.info(f"Applying rule actions to item {queue_item.id}: {actions}")
        rule_engine.apply_actions(queue_item, actions)
        
        # 승인 상태 재평가 (규칙 엔진이 변경했을 수 있음)
        # 하지만 Auto-Approve 로직이 우선이라면? 사용자가 명시적으로 체크해제했으면 승인됨.
        if not item_data.approval_required:
             if queue_item.approval_status != "REJECTED":
                 queue_item.approval_status = "AUTO_APPROVED"

    db.commit()
    db.refresh(queue_item)
    
    # [Auto-Upload Trigger]
    # Only trigger if Auto-Approved AND NOT Scheduled
    if queue_item.approval_status == "AUTO_APPROVED" and queue_item.status == "QUEUED":
        logger.info(f"[FALLBACK] Auto-Approved item {queue_item.id}. Queuing for upload...")
        native_worker.add_task(queue_item.id)
    
    return queue_item


@router.get("/items/{item_id}")
def get_queue_item(item_id: int, db: Session = Depends(get_db)):
    """작업 대기열 항목 상세 조회"""
    item = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Queue item not found")
    return item


@router.patch("/items/{item_id}")
def update_queue_item(
    item_id: int,
    update_data: WorkQueueItemUpdate,
    db: Session = Depends(get_db)
):
    """작업 대기열 항목 수정"""
    item = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    # 업데이트
    for key, value in update_data.dict(exclude_unset=True).items():
        setattr(item, key, value)
    
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(item)
    
    return item


@router.delete("/items/{item_id}")
def delete_queue_item(item_id: int, db: Session = Depends(get_db)):
    """작업 대기열 항목 삭제"""
    item = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    db.delete(item)
    db.commit()
    
    return {"message": "Queue item deleted"}


@router.post("/items/{item_id}/approve")
def approve_queue_item(
    item_id: int,
    approval_data: Optional[ExpertApprovalRequest] = None,
    db: Session = Depends(get_db)
):
    """작업 대기열 항목 승인 (Expert Intervention 지원)"""
    item = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    # [EXPERT INTERVENTION] Apply human edits if provided
    if approval_data:
        if approval_data.script:
            item.description = approval_data.script
            logger.info(f"✍️ [Expert] Script modified for item {item_id}")
            
        if approval_data.instructions:
            # Store instructions in platform_configs for mission context
            if not item.platform_configs:
                item.platform_configs = {}
            item.platform_configs["expert_instructions"] = approval_data.instructions
            logger.info(f"🧠 [Expert] Instructions injected for item {item_id}: {approval_data.instructions}")

            # [EVOLUTION] Persistence to Master Identity
            if approval_data.update_master_identity:
                yt_config = item.platform_configs.get('youtube', {})
                channel_id_field = yt_config.get('channel_id')
                if channel_id_field:
                    channel = db.query(models.BrandChannel).filter(models.BrandChannel.channel_id == channel_id_field).first()
                    if channel:
                        if not channel.expert_identity:
                            channel.expert_identity = {}
                        channel.expert_identity["latest_instructions"] = approval_data.instructions
                        channel.identity_version += 1
                        logger.info(f"🧬 [Evolution] Master Identity updated for channel {channel_id_field}")

    item.approval_status = "APPROVED"
    item.approved_by = approval_data.approved_by if approval_data else "system"
    item.approved_at = datetime.now()
    item.updated_at = datetime.now()
    item.status = "QUEUED"
    
    db.commit()
    db.refresh(item)
    
    # [Native Queue Trigger]
    native_worker.add_task(item.id)
    
    return {
        "status": "APPROVED",
        "item_id": item.id,
        "mode": "expert_intervention" if approval_data else "native_queue"
    }


@router.post("/items/{item_id}/reject")
def reject_queue_item(
    item_id: int,
    reason: str,
    db: Session = Depends(get_db)
):
    """작업 대기열 항목 반려"""
    item = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    item.approval_status = "REJECTED"
    item.rejection_reason = reason
    item.updated_at = datetime.now()
    
    db.commit()
    db.refresh(item)
    return item


@router.post("/batch/shield", response_model=dict)
def batch_apply_shield(
    req: BatchShieldConfigRequest,
    db: Session = Depends(get_db)
):
    from sqlalchemy.orm.attributes import flag_modified
    items = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id.in_(req.item_ids)).all()
    count = 0
    for item in items:
        configs = item.platform_configs or {}
        if 'youtube' not in configs:
            configs['youtube'] = {}
            
        configs['youtube']['anti_association'] = req.shield_configs.get('anti_association', {})
        if 'headless_mode' in req.shield_configs:
            configs['youtube']['headless_mode'] = req.shield_configs['headless_mode']
            
        item.platform_configs = configs
        flag_modified(item, "platform_configs")
        count += 1
    
    db.commit()
    return {"status": "success", "updated": count}

# ============================================
# Draft & Bulk Import AK Endpoints
# ============================================

@router.post("/items/draft", response_model=WorkQueueItemResponse)
def create_draft_item(
    item_data: DraftItemCreate,
    db: Session = Depends(get_db)
):
    """임시 등록 (Draft) - 제목+메타데이터만 저장, 영상은 나중에 첨부"""
    import uuid
    
    queue_item = models.WorkQueueItem(
        title=item_data.title,
        description=item_data.description,
        hashtags=item_data.hashtags,
        tags=item_data.tags,
        source_type=item_data.source_type or "BULK_IMPORT",
        upload_method=item_data.upload_method or "BROWSER_AUTO",
        target_platforms=item_data.target_platforms or ["youtube"],
        platform_configs=item_data.platform_configs or {},
        scheduled_upload_time=item_data.scheduled_upload_time,
        source_batch_id=item_data.source_batch_id or str(uuid.uuid4()),
        source_external_id=item_data.source_external_id,
        source_metadata=item_data.source_metadata or {},
        status="DRAFT",
        approval_status="PENDING",
        upload_progress=0,
        created_at=datetime.now()
    )
    db.add(queue_item)
    db.commit()
    db.refresh(queue_item)
    return queue_item


@router.post("/items/bulk/import", response_model=dict)
def bulk_import(
    data: BulkImportRequest,
    db: Session = Depends(get_db)
):
    """Bulk import (CSV JSON converted) các draft items"""
    import uuid as _uuid
    batch_id = data.source_batch_id or str(_uuid.uuid4())
    created = []
    
    for item_data in data.items:
        queue_item = models.WorkQueueItem(
            title=item_data.title,
            description=item_data.description,
            hashtags=item_data.hashtags,
            tags=item_data.tags,
            source_type=item_data.source_type or "BULK_IMPORT",
            upload_method=item_data.upload_method or "BROWSER_AUTO",
            target_platforms=item_data.target_platforms or ["youtube"],
            platform_configs=item_data.platform_configs or {},
            scheduled_upload_time=item_data.scheduled_upload_time,
            source_batch_id=batch_id,
            source_external_id=item_data.source_external_id,
            source_metadata=item_data.source_metadata or {},
            status="DRAFT",
            approval_status="PENDING",
            upload_progress=0,
            created_at=datetime.now()
        )
        db.add(queue_item)
        created.append({'title': queue_item.title, 'external_id': item_data.source_external_id, 'id': None})
    
    db.commit()
    
    # Refresh for IDs -- first lookup by external_id, fallback to title
    for item in db.query(models.WorkQueueItem).filter(
        models.WorkQueueItem.source_batch_id == batch_id
    ).all():
        for c in created:
            if c.get('external_id') and item.source_external_id == c['external_id']:
                c['id'] = item.id
            elif item.title == c['title']:
                c['id'] = item.id
    
    return {
        "batch_id": batch_id,
        "count": len(created),
        "items": created
    }

@router.patch("/items/{item_id}/attach", response_model=WorkQueueItemResponse)
def attach_video(
    item_id: int,
    data: AttachVideoRequest,
    db: Session = Depends(get_db)
):
    """Draft 초건에 영상 경로를 첨부하고 상태를 PENDING으로 전환
    - item_id 또는 data.source_external_id로 Draft 찾기
    """
    item = None
    if data.source_external_id:
        item = db.query(models.WorkQueueItem).filter(
            models.WorkQueueItem.source_external_id == data.source_external_id,
            models.WorkQueueItem.status == "DRAFT"
        ).first()
    if not item:
        item = db.query(models.WorkQueueItem).filter(
            models.WorkQueueItem.id == item_id,
            models.WorkQueueItem.status == "DRAFT"
        ).first()
    if not item:
        raise HTTPException(404, "Draft not found")
    if not os.path.exists(data.video_file_path):
        raise HTTPException(400, f"Video file not found: {data.video_file_path}")

    item.video_file_path = data.video_file_path
    item.status = "PENDING"
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(item)
    return item

@router.patch("/items/{item_id}/finalize", response_model=dict)
def finalize_draft(
    item_id: int,
    data: FinalizeRequest = None,
    db: Session = Depends(get_db)
):
    """Finalize: PENDING -> AUTO_APPROVED (if approval_required=False) -> QUEUED (trigger upload)"""
    item = db.query(models.WorkQueueItem).filter(
        models.WorkQueueItem.id == item_id,
        models.WorkQueueItem.status.in_(["DRAFT", "PENDING"])
    ).first()
    if not item:
        raise HTTPException(404, "Draft verktding not found")

    if not item.video_file_path:
        raise HTTPException(400, "Video file path not attached yet")

    approval_recomm = (data.approval_required if data else False)
    item.approval_required = approval_recomm
    item.approval_status = "PENDING" if approval_recomm else "AUTO_APPROVED"
    item.status = "QUEUED"
    if data and data.scheduled_upload_time:
        item.scheduled_upload_time = data.scheduled_upload_time
        item.status = "SCHEDULED_UPLOAD"
    if data and data.upload_method:
        item.upload_method = data.upload_method
    if data and data.target_platforms:
        item.target_platforms = data.target_platforms
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(item)

    if item.approval_status == "AUTO_APPROVED":
        from app.services.native_queue_worker import add_task
        add_task(item.id)

    return {
        "item_id": item.id,
        "status": item.status,
        "approval_status": item.approval_status,
        "upload_queued": item.approval_status == "AUTO_APPROVED"
    }

@router.post("/batch/attach", response_model=dict)
def batch_attach_videos(
    data: BulkImportRequest,
    db: Session = Depends(get_db)
):
    """Batch로 external_id 기준 영상 경로 일괄 첨부 (JSON 배열)"""
    results = []
    for item_data in data.items:
        if not item_data.source_external_id:
            results.append({"external_id": None, "status": "skipped", "reason": "no external_id"})
            continue
        draft = db.query(models.WorkQueueItem).filter(
            models.WorkQueueItem.source_external_id == item_data.source_external_id,
            models.WorkQueueItem.status == "DRAFT"
        ).first()
        if not draft:
            results.append({"external_id": item_data.source_external_id, "status": "not_found"})
            continue
        video_path = item_data.source_metadata.get("video_file_path") if item_data.source_metadata else None
        if not video_path or not os.path.exists(video_path):
            results.append({"external_id": item_data.source_external_id, "status": "no_video", "item_id": draft.id})
            continue
        draft.video_file_path = video_path
        draft.status = "PENDING"
        draft.updated_at = datetime.now()
        results.append({"external_id": item_data.source_external_id, "item_id": draft.id, "status": "attached"})

    db.commit()
    return {"results": results}


@router.post("/batch/finalize", response_model=dict)
def batch_finalize_drafts(
    data: BulkImportRequest,
    db: Session = Depends(get_db)
):
    """Batch finalize -- external_id 기준 DRAFT/PENDING 아이템을 QUEUED로 일괄 전환"""
    finalized = []
    for item_data in data.items:
        ext_id = item_data.source_external_id
        item = None
        if ext_id:
            item = db.query(models.WorkQueueItem).filter(
                models.WorkQueueItem.source_external_id == ext_id,
                models.WorkQueueItem.status.in_(["DRAFT", "PENDING"])
            ).first()
        if not item:
            finalized.append({"external_id": ext_id, "status": "not_found"})
            continue
        if not item.video_file_path:
            finalized.append({"external_id": ext_id, "item_id": item.id, "status": "no_video"})
            continue
        item.approval_required = False
        item.approval_status = "AUTO_APPROVED"
        item.status = "QUEUED"
        item.updated_at = datetime.now()
        from app.services.native_queue_worker import add_task
        add_task(item.id)
        finalized.append({"external_id": ext_id, "item_id": item.id, "status": "finalized"})

    db.commit()
    return {"count": len(finalized), "results": finalized}


class BulkFileUploadRequest(BaseModel):
    base64_file: str  # base64-encoded file content
    file_name: str
    source_batch_id: Optional[str] = None
    default_upload_method: Optional[str] = "BROWSER_AUTO"
    default_target_platforms: Optional[List[str]] = ["youtube"]


@router.post("/bulk/upload-file", response_model=dict)
def bulk_upload_file(
    data: BulkFileUploadRequest,
    db: Session = Depends(get_db)
):
    """CSV (.csv) or Excel (.xlsx) 파일을 업로드해서 자동으로 Draft 임시 등록
    - CSV: comma separated, headers on row 1
    - Excel: .xlsx only, first sheet, headers on row 1
    - external_id를 filename+row로 자동 생성 (override 가능)
    - 각 행은 제목(title) + 설명(description) + external_id를 JSON으로 변환
    """
    import base64
    import io
    import uuid as _uuid

    raw = base64.b64decode(data.base64_file)
    ext = os.path.splitext(data.file_name or "")[1].lower() if hasattr(data, 'file_name') else ".csv"
    if not hasattr(data, 'file_name'):
        ext = ".csv"

    rows = []
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(400, "openpyxl not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [cell.strip() if isinstance(cell, str) else str(cell or "").strip() for cell in row]
                continue
            vals = [cell if isinstance(cell, str) else (str(cell) if cell is not None else "") for cell in row]
            row_dict = {}
            for j, h in enumerate(headers):
                row_dict[h] = vals[j] if j < len(vals) else ""
            rows.append(row_dict)
        wb.close()
    else:
        text = raw.decode("utf-8-sig")
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if not lines:
            raise HTTPException(400, "Empty CSV file")
        headers = [h.strip() for h in lines[0].split(",")]
        for line in lines[1:]:
            vals = [v.strip() for v in line.split(",")]
            row_dict = {}
            for j, h in enumerate(headers):
                row_dict[h] = vals[j] if j < len(vals) else ""
            rows.append(row_dict)

    if not rows:
        raise HTTPException(400, "No data rows found")

    batch_id = data.source_batch_id or str(_uuid.uuid4())
    created = []

    title_col = next((h for h in headers if h.lower() in ("title", "제목", "name")), None)
    desc_col = next((h for h in headers if h.lower() in ("description", "desc", "설명")), None)
    ext_col = next((h for h in headers if h.lower() in ("external_id", "외부id", "id")), None)
    hashtag_col = next((h for h in headers if h.lower() in ("hashtags",)), None)
    tag_col = next((h for h in headers if h.lower() in ("tags", "태그")), None)
    um_col = next((h for h in headers if h.lower() in ("upload_method", "업로드방식")), None)
    plat_col = next((h for h in headers if h.lower() in ("platforms", "플랫폼")), None)
    pp_col = next((h for h in headers if h.lower() in ("platform_privacy", "공개설정")), None)
    st_col = next((h for h in headers if h.lower() in ("scheduled_time", "예약시간")), None)

    for row in rows:
        title = row.get(title_col, "") if title_col else f"Item {len(created) + 1}"
        description = row.get(desc_col, "") if desc_col else ""
        external_id = row.get(ext_col, "") if ext_col else f"{batch_id}_{len(created) + 1:04d}"

        hashtags_raw = row.get(hashtag_col, "") if hashtag_col else ""
        tags_raw = row.get(tag_col, "") if tag_col else ""
        parsed_hashtags = [t if t.startswith('#') else f"#{t}" for t in hashtags_raw.split() if t.strip()] if hashtags_raw else None
        parsed_tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else None

        upload_method = row.get(um_col, "").strip() if um_col else None
        if not upload_method:
            upload_method = data.default_upload_method or "BROWSER_AUTO"

        platforms_raw = row.get(plat_col, "").strip() if plat_col else ""
        target_platforms = [p.strip() for p in platforms_raw.split(",") if p.strip()] if platforms_raw else (data.default_target_platforms or ["youtube"])

        privacy = row.get(pp_col, "").strip().lower() if pp_col else None
        platform_configs = {}
        if privacy and target_platforms:
            for p in target_platforms:
                platform_configs[p] = {"privacy": privacy}

        scheduled_raw = row.get(st_col, "").strip() if st_col else ""
        scheduled_upload_time = None
        if scheduled_raw:
            try:
                scheduled_upload_time = datetime.fromisoformat(scheduled_raw)
            except ValueError:
                pass

        queue_item = models.WorkQueueItem(
            title=title.strip() or f"Item {len(created) + 1}",
            description=description.strip(),
            hashtags=parsed_hashtags,
            tags=parsed_tags,
            source_type="BULK_IMPORT",
            upload_method=upload_method or "BROWSER_AUTO",
            target_platforms=target_platforms,
            platform_configs=platform_configs,
            scheduled_upload_time=scheduled_upload_time,
            source_batch_id=batch_id,
            source_external_id=external_id,
            source_metadata={"original_row": row},
            status="DRAFT",
            approval_status="PENDING",
            upload_progress=0,
            created_at=datetime.now()
        )
        db.add(queue_item)
        db.flush()
        created.append({
            "id": queue_item.id,
            "external_id": external_id,
            "title": queue_item.title,
            "status": "DRAFT"
        })

    db.commit()
    return {
        "batch_id": batch_id,
        "count": len(created),
        "items": created
    }

# === Template Download Endpoints ===

TEMPLATE_COLUMNS = [
    "title", "description", "hashtags", "tags",
    "external_id", "upload_method", "platforms", "platform_privacy", "scheduled_time"
]

TEMPLATE_SAMPLE_ROWS = [
    ["재미있는 고양이 영상", "고양이가 장난감과 노는 모습을 담은 영상입니다", "#cat #funny", "cat,funny", "cat_001", "BROWSER_AUTO", "youtube", "private", ""],
    ["하늘 풍경 타임랩스", "아름다운 노을과 구름의 변화를 담았습니다", "#sky #timelapse", "sky, timelapse, nature", "sky_002", "API", "youtube,tiktok", "unlisted", "2026-08-01 09:00"],
]

import csv
import io
import tempfile
from fastapi.responses import StreamingResponse, FileResponse

HAS_OPENPYXL = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass

def _generate_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(TEMPLATE_COLUMNS)
    for row in TEMPLATE_SAMPLE_ROWS:
        writer.writerow(row)
    return output.getvalue().encode("utf-8-sig")

@router.get("/template/csv")
def download_template_csv():
    content = _generate_template_csv()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=VLStudio_work_queue_template.csv"}
    )

@router.get("/template/xlsx")
def download_template_xlsx():
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=501, detail="openpyxl is not installed. Install with: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work Queue Template"
    ws.append(TEMPLATE_COLUMNS)
    for row in TEMPLATE_SAMPLE_ROWS:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(tmp.name)
        tmp.close()
        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="VLStudio_work_queue_template.xlsx",
            background=None,
        )
    except Exception:
        os.unlink(tmp.name)
        raise


@router.get("/stats")
def get_queue_stats(db: Session = Depends(get_db)):
    """작업 대기열 통계"""
    draft = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "DRAFT").count()
    pending = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "PENDING").count()
    total = db.query(models.WorkQueueItem).count()
    queued = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "QUEUED").count()
    uploading = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "UPLOADING").count()
    completed = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "COMPLETED").count()
    failed = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "FAILED").count()
    verifying = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "VERIFYING").count()
    failed_review = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.status == "FAILED_REVIEW").count()
    
    pending_approval = db.query(models.WorkQueueItem).filter(
        models.WorkQueueItem.approval_status == "PENDING"
    ).count()
    
    return {
        "total": total,
        "draft": draft,
        "pending": pending,
        "queued": queued,
        "uploading": uploading,
        "completed": completed,
        "failed": failed,
        "verifying": verifying,
        "failed_review": failed_review,
        "pending_approval": pending_approval
    }

# === WebSocket for Real-time Progress ===

class ConnectionManager:
    def __init__(self):
        # item_id -> List[WebSocket]
        self.active_connections: dict[int, List[WebSocket]] = {}

    async def connect(self, websocket: WebSocket, item_id: int):
        await websocket.accept()
        if item_id not in self.active_connections:
            self.active_connections[item_id] = []
        self.active_connections[item_id].append(websocket)
        logger.info(f"[OK] WebSocket Client connected to item {item_id}")

    def disconnect(self, websocket: WebSocket, item_id: int):
        if item_id in self.active_connections:
            if websocket in self.active_connections[item_id]:
                self.active_connections[item_id].remove(websocket)
            if not self.active_connections[item_id]:
                del self.active_connections[item_id]
        logger.info(f"🔌 WebSocket Client disconnected from item {item_id}")

    async def broadcast(self, item_id: int, message: dict):
        if item_id in self.active_connections:
            # Copy list to avoid modification during iteration
            for connection in self.active_connections[item_id][:]:
                try:
                    await connection.send_json(message)
                except Exception as e:
                    logger.warning(f"Failed to send to socket: {e}")
                    self.disconnect(connection, item_id)

manager = ConnectionManager()

@router.websocket("/ws/progress/{item_id}")
async def websocket_endpoint(websocket: WebSocket, item_id: int):
    # [Security] Validate item exists?
    # For now, just accept to avoid 403 loop if DB is locked
    await manager.connect(websocket, item_id)
    try:
        while True:
            # Just keep connection open, maybe handle pings
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket, item_id)
    except Exception as e:
        logger.error(f"WebSocket Error: {e}")
        manager.disconnect(websocket, item_id)

# Expose broadcast for worker
async def notify_progress(item_id: int, status: str, progress: int, log: str = None):
    msg = {
        "item_id": item_id,
        "status": status,
        "progress": progress,
        "log": log,
        "timestamp": datetime.now().isoformat()
    }
    await manager.broadcast(item_id, msg)

@router.post("/items/{item_id}/upload")
def trigger_upload(item_id: int, db: Session = Depends(get_db)):
    """
    업로드 작업 수동 트리거 (Native Queue)
    """
    item = db.query(models.WorkQueueItem).filter(models.WorkQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    # 승인 확인
    if item.approval_status not in ["APPROVED", "AUTO_APPROVED"]:
        raise HTTPException(400, "Item not approved")
    
    # Native Queue Trigger
    native_worker.add_task(item.id)
    
    return {
        "message": "Upload task triggered",
        "item_id": item_id,
        "mode": "native_queue"
    }


# ... (Batch Classes) ...

@router.post("/batch/approve")
def batch_approve(
    request: BatchApproveRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """일괄 승인 (Native Queue)"""
    approved_items = []
    failed_items = []
    items_to_queue = []
    
    for item_id in request.item_ids:
        try:
            item = db.query(models.WorkQueueItem).filter(
                models.WorkQueueItem.id == item_id
            ).first()
            
            if not item:
                failed_items.append({"item_id": item_id, "reason": "Not found"})
                continue
            
            item.approval_status = "APPROVED"
            item.approved_by = request.approved_by
            item.approved_at = datetime.now()
            item.updated_at = datetime.now()
            item.status = "QUEUED"
            
            items_to_queue.append(item.id)
            
            approved_items.append({
                "item_id": item.id,
                "mode": "native_queue"
            })
                
        except Exception as e:
            logger.error(f"Failed to approve item {item_id}: {e}")
            failed_items.append({"item_id": item_id, "reason": str(e)})
    
    db.commit()
    
    # [Native Queue Trigger] after commit
    for item_id in items_to_queue:
        native_worker.add_task(item_id)
    
    return {
        "approved": len(approved_items),
        "failed": len(failed_items),
        "approved_items": approved_items,
        "failed_items": failed_items
    }


@router.post("/batch/reject")
def batch_reject(
    request: BatchRejectRequest,
    db: Session = Depends(get_db)
):
    """일괄 반려"""
    rejected_items = []
    failed_items = []
    
    for item_id in request.item_ids:
        try:
            item = db.query(models.WorkQueueItem).filter(
                models.WorkQueueItem.id == item_id
            ).first()
            
            if not item:
                failed_items.append({"item_id": item_id, "reason": "Not found"})
                continue
            
            item.approval_status = "REJECTED"
            item.rejection_reason = request.reason
            item.updated_at = datetime.now()
            
            rejected_items.append(item.id)
            
        except Exception as e:
            logger.error(f"Failed to reject item {item_id}: {e}")
            failed_items.append({"item_id": item_id, "reason": str(e)})
    
    db.commit()
    
    return {
        "rejected": len(rejected_items),
        "failed": len(failed_items),
        "rejected_items": rejected_items,
        "failed_items": failed_items
    }


@router.post("/batch/reset")
def batch_reset(
    request: BatchResetRequest,
    db: Session = Depends(get_db)
):
    """
    일괄 상태 초기화 (실패한 항목 재시도용)
    Status -> QUEUED
    Approval -> PENDING
    Failure Reason -> None
    """
    reset_items = []
    failed_items = []
    
    for item_id in request.item_ids:
        try:
            item = db.query(models.WorkQueueItem).filter(
                models.WorkQueueItem.id == item_id
            ).first()
            
            if not item:
                failed_items.append({"item_id": item_id, "reason": "Not found"})
                continue
            
            # Reset Status
            item.status = "QUEUED"
            item.approval_status = "PENDING"
            item.failure_reason = None
            item.upload_progress = 0
            item.updated_at = datetime.now()
            
            reset_items.append(item.id)
            
        except Exception as e:
            logger.error(f"Failed to reset item {item_id}: {e}")
            failed_items.append({"item_id": item_id, "reason": str(e)})
    
    db.commit()
    
    return {
        "reset": len(reset_items),
        "failed": len(failed_items),
        "reset_items": reset_items,
        "failed_items": failed_items
    }



@router.post("/batch/delete")
def batch_delete(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db)
):
    """일괄 삭제"""
    deleted_items = []
    failed_items = []
    
    for item_id in request.item_ids:
        try:
            item = db.query(models.WorkQueueItem).filter(
                models.WorkQueueItem.id == item_id
            ).first()
            
            if not item:
                failed_items.append({"item_id": item_id, "reason": "Not found"})
                continue
            
            db.delete(item)
            deleted_items.append(item_id)
            
        except Exception as e:
            logger.error(f"Failed to delete item {item_id}: {e}")
            failed_items.append({"item_id": item_id, "reason": str(e)})
    
    db.commit()
    
    return {
        "deleted": len(deleted_items),
        "failed": len(failed_items),
        "deleted_items": deleted_items,
        "failed_items": failed_items
    }


@router.post("/priority/update")
def update_priority(
    request: PriorityUpdateRequest,
    db: Session = Depends(get_db)
):
    """우선순위 업데이트"""
    item = db.query(models.WorkQueueItem).filter(
        models.WorkQueueItem.id == request.item_id
    ).first()
    
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    item.upload_priority = request.priority
    item.updated_at = datetime.now()
    
    db.commit()
    db.refresh(item)
    
    return item


@router.post("/priority/reorder")
def reorder_priorities(
    item_ids: List[int],
    db: Session = Depends(get_db)
):
    """
    우선순위 재정렬 (드래그 앤 드롭)
    
    Args:
        item_ids: 새로운 순서대로 정렬된 항목 ID 리스트
    """
    # 우선순위를 역순으로 할당 (첫 번째 항목이 가장 높은 우선순위)
    for index, item_id in enumerate(item_ids):
        item = db.query(models.WorkQueueItem).filter(
            models.WorkQueueItem.id == item_id
        ).first()
        
        if item:
            # 우선순위: 리스트 길이 - 인덱스
            item.upload_priority = len(item_ids) - index
            item.updated_at = datetime.now()
    
    db.commit()
    
    return {"message": "Priorities reordered", "count": len(item_ids)}


@router.post("/schedule/delay")
def schedule_delayed_upload(
    item_id: int,
    delay_minutes: int,
    db: Session = Depends(get_db)
):
    """
    지연 업로드 스케줄링
    
    Args:
        item_id: 항목 ID
        delay_minutes: 지연 시간 (분)
    """
    item = db.query(models.WorkQueueItem).filter(
        models.WorkQueueItem.id == item_id
    ).first()
    
    if not item:
        raise HTTPException(404, "Queue item not found")
    
    item.upload_delay_minutes = delay_minutes
    item.updated_at = datetime.now()
    
    db.commit()
    db.refresh(item)
    
    return item


# === AI Metadata Generation ===

class MetadataGenerationRequest(BaseModel):
    video_path: str
    platform: str = "youtube"  # youtube, tiktok, instagram


@router.post("/generate-metadata")
def generate_metadata(
    request: MetadataGenerationRequest,
    db: Session = Depends(get_db)
):
    """
    AI 기반 메타데이터 자동 생성
    
    영상 파일에서 자막을 추출하고 AI를 사용하여 플랫폼별로 최적화된
    제목, 설명, 해시태그를 자동으로 생성합니다.
    
    Args:
        request: 영상 경로 및 플랫폼 정보
        
    Returns:
        생성된 메타데이터 (title, description/caption, hashtags)
    """
    try:
        # 파일 존재 확인
        if not os.path.exists(request.video_path):
            # [FIX] Try to find in downloads folder if relative path provided
            from app import crud
            from app.config import settings as settings_conf
            settings = crud.get_settings(db)
            root_path = settings.root_download_path if settings and settings.root_download_path else settings_conf.MEDIA_ROOT
            download_path = os.path.join(root_path, request.video_path)
            
            if os.path.exists(download_path):
                logger.info(f"files resolved to: {download_path}")
                request.video_path = download_path
            else:
                logger.error(f"Video file not found at: {request.video_path} OR {download_path}")
                # [FIX] Return 400 (Bad Request) instead of 404 to distinguish from 'Route Not Found'
                # Also provide a helpful hint to the user.
                from app.config import settings
                import platform
                example_path = "F:\\Videos\\file.mp4" if platform.system() == "Windows" else "/home/user/viral_loop_media/videos/file.mp4"
                raise HTTPException(
                    status_code=400, 
                    detail=f"Video file not found. Please provide the FULL ABSOLUTE PATH (e.g., {example_path}). System checked: '{request.video_path}' and '{download_path}'"
                )
        
        # AI 메타데이터 서비스 사용
        from app.services.ai_metadata_service import AIMetadataService
        
        service = AIMetadataService(db)
        metadata = service.generate_metadata(
            video_path=request.video_path,
            platform=request.platform.lower()
        )
        
        logger.info(f"[OK] Generated metadata for {request.platform}: {metadata.get('title', 'N/A')}")
        
        return {
            "success": True,
            "platform": request.platform,
            "metadata": metadata
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Metadata generation failed: {e}")
        raise HTTPException(500, f"Failed to generate metadata: {str(e)}")


# === Video Streaming ===

@router.get("/stream")
def stream_video(path: str):
    """
    Local Video Streaming for Work Queue
    Allows playing files from absolute paths (e.g. F:/...)
    """
    if not os.path.exists(path):
        raise HTTPException(404, "File not found")
        
    from fastapi.responses import FileResponse
    return FileResponse(path, media_type="video/mp4")
